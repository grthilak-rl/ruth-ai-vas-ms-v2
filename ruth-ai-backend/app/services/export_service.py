"""Export service for analytics data.

Provides CSV, XLSX, and PDF export functionality for violations data.
Per analytics-design.md Section 7 (Export Data Flow).
"""

import csv
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import and_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.logging import get_logger
from app.models import Device, Violation

logger = get_logger(__name__)


class ExportService:
    """Service for exporting analytics data in various formats."""

    def __init__(self, db: AsyncSession):
        """Initialize export service.

        Args:
            db: Database session
        """
        self.db = db

    async def export_violations(
        self,
        format: str,
        from_time: datetime,
        to_time: datetime,
        camera_ids: list[str] | None = None,
        violation_types: list[str] | None = None,
        statuses: list[str] | None = None,
        include_headers: bool = True,
        include_timestamps: bool = True,
        include_raw_confidence: bool = False,
        include_evidence_urls: bool = False,
        include_bounding_boxes: bool = False,
    ) -> tuple[bytes, str, str]:
        """Export violations data.

        Args:
            format: Export format (csv, xlsx, pdf)
            from_time: Start of time range
            to_time: End of time range
            camera_ids: Optional camera filter
            violation_types: Optional violation type filter
            statuses: Optional status filter
            include_headers: Include column headers
            include_timestamps: Include ISO 8601 timestamps
            include_raw_confidence: Include raw confidence scores
            include_evidence_urls: Include evidence URLs
            include_bounding_boxes: Include bounding box coordinates

        Returns:
            Tuple of (file_bytes, content_type, filename)
        """
        # Fetch violations data
        violations = await self._fetch_violations(
            from_time,
            to_time,
            camera_ids,
            violation_types,
            statuses,
        )

        logger.info(
            "Exporting violations",
            format=format,
            count=len(violations),
            from_time=from_time.isoformat(),
            to_time=to_time.isoformat(),
        )

        # Generate export based on format
        if format == "csv":
            return await self._export_csv(
                violations,
                from_time,
                to_time,
                include_headers,
                include_timestamps,
                include_raw_confidence,
                include_evidence_urls,
                include_bounding_boxes,
            )
        elif format == "xlsx":
            return await self._export_xlsx(
                violations,
                from_time,
                to_time,
                include_raw_confidence,
            )
        elif format == "pdf":
            return await self._export_pdf(
                violations,
                from_time,
                to_time,
            )
        else:
            raise ValueError(f"Unsupported export format: {format}")

    async def _fetch_violations(
        self,
        from_time: datetime,
        to_time: datetime,
        camera_ids: list[str] | None = None,
        violation_types: list[str] | None = None,
        statuses: list[str] | None = None,
    ) -> list[Violation]:
        """Fetch violations matching filters."""
        filters = [
            Violation.timestamp >= from_time,
            Violation.timestamp < to_time,
        ]

        if camera_ids:
            filters.append(Violation.device_id.in_(camera_ids))

        if violation_types:
            filters.append(Violation.type.in_(violation_types))

        if statuses:
            filters.append(Violation.status.in_(statuses))

        stmt = (
            select(Violation)
            .where(and_(*filters))
            .order_by(Violation.timestamp.desc())
        )

        result = await self.db.execute(stmt)
        return list(result.scalars().all())

    async def _export_csv(
        self,
        violations: list[Violation],
        from_time: datetime,
        to_time: datetime,
        include_headers: bool,
        include_timestamps: bool,
        include_raw_confidence: bool,
        include_evidence_urls: bool,
        include_bounding_boxes: bool,
    ) -> tuple[bytes, str, str]:
        """Export violations as CSV."""
        output = StringIO()
        writer = csv.writer(output)

        # Define columns
        columns = ["id", "type", "camera_name", "status"]

        if include_raw_confidence:
            columns.append("confidence_raw")
        else:
            columns.append("confidence_category")

        if include_timestamps:
            columns.extend(["timestamp", "created_at"])

        columns.extend(["reviewed_by", "reviewed_at"])

        if include_evidence_urls:
            columns.extend(["snapshot_url", "bookmark_url"])

        if include_bounding_boxes:
            columns.append("bounding_boxes")

        columns.extend(["model_id", "model_version"])

        # Write header
        if include_headers:
            writer.writerow(columns)

        # Write data
        for v in violations:
            row = [
                str(v.id),
                v.type.value,
                v.camera_name,
                v.status.value,
            ]

            # Confidence
            if include_raw_confidence:
                row.append(f"{v.confidence:.3f}")
            else:
                if v.confidence >= 0.8:
                    row.append("High")
                elif v.confidence >= 0.6:
                    row.append("Medium")
                else:
                    row.append("Low")

            # Timestamps
            if include_timestamps:
                row.extend([
                    v.timestamp.isoformat(),
                    v.created_at.isoformat(),
                ])

            # Review info
            row.extend([
                v.reviewed_by or "",
                v.reviewed_at.isoformat() if v.reviewed_at else "",
            ])

            # Evidence URLs (placeholder - would need to construct from evidence table)
            if include_evidence_urls:
                row.extend(["", ""])  # TODO: Fetch from evidence table if needed

            # Bounding boxes
            if include_bounding_boxes:
                import json
                row.append(json.dumps(v.bounding_boxes) if v.bounding_boxes else "")

            # Model info
            row.extend([v.model_id, v.model_version])

            writer.writerow(row)

        # Convert to bytes
        csv_bytes = output.getvalue().encode("utf-8")

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        filename = f"ruth-ai-analytics-{timestamp}.csv"

        return csv_bytes, "text/csv", filename

    async def _export_xlsx(
        self,
        violations: list[Violation],
        from_time: datetime,
        to_time: datetime,
        include_raw_confidence: bool,
    ) -> tuple[bytes, str, str]:
        """Export violations as XLSX with formatting."""
        wb = Workbook()

        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Add summary information
        ws_summary["A1"] = "Ruth AI Analytics Report"
        ws_summary["A1"].font = Font(size=16, bold=True)

        ws_summary["A3"] = "Time Range:"
        ws_summary["B3"] = f"{from_time.strftime('%Y-%m-%d %H:%M')} to {to_time.strftime('%Y-%m-%d %H:%M')}"

        ws_summary["A4"] = "Total Violations:"
        ws_summary["B4"] = len(violations)

        ws_summary["A5"] = "Generated:"
        ws_summary["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Status breakdown
        from collections import Counter
        status_counts = Counter(v.status.value for v in violations)

        row = 7
        ws_summary["A7"] = "Status Breakdown:"
        ws_summary["A7"].font = Font(bold=True)
        for status, count in status_counts.items():
            row += 1
            ws_summary[f"A{row}"] = status.title()
            ws_summary[f"B{row}"] = count

        # Violations List Sheet
        ws_violations = wb.create_sheet("Violations")

        # Headers
        headers = ["ID", "Type", "Camera", "Status", "Confidence", "Timestamp", "Reviewed By", "Reviewed At"]
        ws_violations.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws_violations.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for v in violations:
            if include_raw_confidence:
                confidence_display = f"{v.confidence:.3f}"
            else:
                if v.confidence >= 0.8:
                    confidence_display = "High"
                elif v.confidence >= 0.6:
                    confidence_display = "Medium"
                else:
                    confidence_display = "Low"

            ws_violations.append([
                str(v.id)[:8] + "...",  # Truncated UUID
                v.type.value.replace("_", " ").title(),
                v.camera_name,
                v.status.value.title(),
                confidence_display,
                v.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                v.reviewed_by or "",
                v.reviewed_at.strftime("%Y-%m-%d %H:%M:%S") if v.reviewed_at else "",
            ])

        # Auto-size columns
        for column in ws_violations.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_violations.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws_violations.freeze_panes = "A2"

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        filename = f"ruth-ai-analytics-{timestamp}.xlsx"

        return output.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename

    async def _export_pdf(
        self,
        violations: list[Violation],
        from_time: datetime,
        to_time: datetime,
    ) -> tuple[bytes, str, str]:
        """Export violations as PDF report."""
        output = BytesIO()

        # Create PDF document
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18,
        )

        # Container for elements
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1,  # Center
        )
        elements.append(Paragraph("Ruth AI Analytics Report", title_style))
        elements.append(Spacer(1, 12))

        # Metadata
        meta_style = styles["Normal"]
        elements.append(Paragraph(f"<b>Time Range:</b> {from_time.strftime('%Y-%m-%d %H:%M')} to {to_time.strftime('%Y-%m-%d %H:%M')}", meta_style))
        elements.append(Paragraph(f"<b>Total Violations:</b> {len(violations)}", meta_style))
        elements.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", meta_style))
        elements.append(Spacer(1, 20))

        # Summary section
        from collections import Counter
        status_counts = Counter(v.status.value for v in violations)
        type_counts = Counter(v.type.value for v in violations)

        elements.append(Paragraph("<b>Executive Summary</b>", styles['Heading2']))
        elements.append(Spacer(1, 12))

        # Status breakdown
        elements.append(Paragraph("<b>By Status:</b>", styles['Heading3']))
        for status, count in status_counts.most_common():
            pct = (count / len(violations) * 100) if violations else 0
            elements.append(Paragraph(f"• {status.title()}: {count} ({pct:.1f}%)", meta_style))
        elements.append(Spacer(1, 12))

        # Type breakdown
        elements.append(Paragraph("<b>By Type:</b>", styles['Heading3']))
        for vtype, count in type_counts.most_common():
            pct = (count / len(violations) * 100) if violations else 0
            elements.append(Paragraph(f"• {vtype.replace('_', ' ').title()}: {count} ({pct:.1f}%)", meta_style))

        elements.append(PageBreak())

        # Violations list
        elements.append(Paragraph("<b>Detailed Violations List</b>", styles['Heading2']))
        elements.append(Spacer(1, 12))

        # Create table
        table_data = [
            ["Type", "Camera", "Status", "Confidence", "Timestamp"]
        ]

        for v in violations[:100]:  # Limit to first 100 for PDF
            if v.confidence >= 0.8:
                confidence_display = "High"
            elif v.confidence >= 0.6:
                confidence_display = "Medium"
            else:
                confidence_display = "Low"

            table_data.append([
                v.type.value.replace("_", " ").title(),
                v.camera_name[:20],  # Truncate long names
                v.status.value.title(),
                confidence_display,
                v.timestamp.strftime("%Y-%m-%d %H:%M"),
            ])

        # Create table with styling
        table = Table(table_data, colWidths=[1.5*inch, 1.5*inch, 1*inch, 1*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        elements.append(table)

        if len(violations) > 100:
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(f"<i>Note: Showing first 100 of {len(violations)} violations</i>", meta_style))

        # Build PDF
        doc.build(elements)

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        filename = f"ruth-ai-analytics-{timestamp}.pdf"

        return output.getvalue(), "application/pdf", filename
